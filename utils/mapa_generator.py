"""
Gerador do Mapa de Consignação — Excel por cliente
Formato fiel ao modelo real: 13.04.26 SALDO CONSIGNADO - L.xls

Estrutura do arquivo gerado (por cliente):
  Linha 0  : totalizadores (Qtde Remet | Dev/Acert | Saldo | Valor Liq | Valor Bruto)
  Linha 1  : cabeçalhos das colunas
  Linhas 2+: dados do saldo daquele cliente
  Coluna extra: "Qtde a Faturar" — campo em branco para o cliente preencher

Enviado por email para o cliente com instrução de retorno.
"""

import io
from datetime import datetime
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, PatternFill, Border, Side,
        numbers as xl_numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ─── PALETA ──────────────────────────────────────────────────────────────────
COR_CABECALHO    = "1E3A5F"   # azul escuro
COR_TOTAIS       = "2D5F8A"   # azul médio
COR_FATURAR      = "1A5E38"   # verde escuro — coluna de ação do cliente
COR_TEXTO_BRANCO = "FFFFFF"
COR_ZEBRA_PAR    = "EBF2FA"   # azul muito claro
COR_ZEBRA_IMPAR  = "FFFFFF"
COR_FATURAR_CELL = "E8F5E9"   # verde clarinho para células da coluna
COR_BORDA        = "B0C4D8"


def _border(style="thin"):
    s = Side(style=style, color=COR_BORDA)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=9):
    return Font(name="Calibri", bold=bold, color=color, size=size)


def gerar_mapa_excel(
    df_saldo: pd.DataFrame,
    codigo_cliente: str,
    razao_social: str,
    cnpj: str = "",
    mes_referencia: str = "",
    codigo_rastreio: str = "",
    vendedor_nome: str = "",
) -> bytes:
    """
    Gera o Excel do mapa de consignação para um cliente específico.

    Parâmetros
    ----------
    df_saldo : DataFrame com os dados do saldo deste cliente
               (colunas esperadas: as do schema do banco)
    codigo_cliente, razao_social, cnpj : identificação do cliente
    mes_referencia : str no formato 'YYYY-MM'
    codigo_rastreio : str gerado pelo banco para rastreamento
    vendedor_nome : nome do vendedor remetente

    Retorna
    -------
    bytes : conteúdo do arquivo .xlsx em memória
    """
    if not OPENPYXL_OK:
        raise ImportError("openpyxl não instalado. Execute: pip install openpyxl")

    # ── 1. Prepara o DataFrame ────────────────────────────────────────────────
    df = df_saldo.copy()

    # Garante tipos numéricos
    for col in ["qtde_remessa", "qtde_dev_acert", "qtde_saldo", "valor_liquido", "valor_bruto"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Ordena por título
    if "titulo" in df.columns:
        df = df.sort_values("titulo")

    # ── 2. Mapeamento de colunas: banco → cabeçalho Excel ────────────────────
    COLUNAS = [
        ("cod_loja",       "Cod+Loja"),
        ("cnpj",           "CNPJ"),
        ("codigo_cliente", "Cod."),
        ("loja",           "Lj."),
        ("cod_gcon",       "Gcon"),
        ("uf",             "UF"),
        ("razao_social",   "Cliente"),
        ("nf_serie",       "NF/Serie"),
        ("data_emissao",   "Data Emissão"),
        ("isbn",           "Codigo"),
        ("cod_barras",     "Cod Barras"),
        ("titulo",         "Titulo"),
        ("autor",          "Autor"),
        ("desconto",       "Desconto"),
        ("status_titulo",  "STATUS"),
        ("qtde_remessa",   "Qtde Remet"),
        ("qtde_dev_acert", "Qtde Dev/Acert"),
        ("qtde_saldo",     "Qtde Saldo"),
        ("valor_liquido",  "Valor Liquido"),
        ("valor_bruto",    "Valor Bruto"),
        # coluna extra — cliente preenche
        (None,             "Qtde a Faturar ←"),
        (None,             "Qtde a Devolver ←"),
        (None,             "Observação"),
    ]

    # Filtra só colunas que existem no df (as de banco)
    cols_excel = []
    for db_col, header in COLUNAS:
        if db_col is None or db_col in df.columns:
            cols_excel.append((db_col, header))

    n_cols_dados = sum(1 for c, _ in cols_excel if c is not None)
    n_cols_total = len(cols_excel)

    # Índice das colunas numéricas de quantidade e valor
    idx_qtde_remet  = next((i for i, (c, _) in enumerate(cols_excel) if c == "qtde_remessa"), None)
    idx_qtde_dev    = next((i for i, (c, _) in enumerate(cols_excel) if c == "qtde_dev_acert"), None)
    idx_qtde_saldo  = next((i for i, (c, _) in enumerate(cols_excel) if c == "qtde_saldo"), None)
    idx_vliq        = next((i for i, (c, _) in enumerate(cols_excel) if c == "valor_liquido"), None)
    idx_vbruto      = next((i for i, (c, _) in enumerate(cols_excel) if c == "valor_bruto"), None)

    # ── 3. Cria o workbook ───────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapa de Consignação"

    # ── 4. Linha de identificação do cliente (linha 1) ───────────────────────
    mes_fmt = _fmt_mes(mes_referencia)
    info_header = (
        f"MAPA DE CONSIGNAÇÃO  |  {razao_social}"
        f"  |  CNPJ: {cnpj or '—'}"
        f"  |  Referência: {mes_fmt}"
        f"  |  Rastreio: {codigo_rastreio}"
        f"  |  Vendedor: {vendedor_nome}"
        f"  |  Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols_total)
    cell_info = ws.cell(row=1, column=1, value=info_header)
    cell_info.font = Font(name="Calibri", bold=True, color=COR_TEXTO_BRANCO, size=9)
    cell_info.fill = _fill(COR_CABECALHO)
    cell_info.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = 18

    # ── 5. Linha de instrução (linha 2) ──────────────────────────────────────
    instrucao = (
        "INSTRUÇÃO: Verifique os títulos em seu estoque, preencha 'Qtde a Faturar' e/ou "
        "'Qtde a Devolver' e retorne este arquivo por email. "
        f"Mantenha o assunto original para rastreamento. Código: {codigo_rastreio}"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols_total)
    cell_instr = ws.cell(row=2, column=1, value=instrucao)
    cell_instr.font = Font(name="Calibri", italic=True, color="1A3A5F", size=9)
    cell_instr.fill = _fill("D6E8F8")
    cell_instr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[2].height = 16

    # ── 6. Linha de totalizadores (linha 3 — formato do arquivo original) ────
    totais_label = "TOTAIS"
    tot_remet  = int(df["qtde_remessa"].sum())  if "qtde_remessa"  in df.columns else 0
    tot_dev    = int(df["qtde_dev_acert"].sum()) if "qtde_dev_acert" in df.columns else 0
    tot_saldo  = int(df["qtde_saldo"].sum())    if "qtde_saldo"    in df.columns else 0
    tot_vliq   = float(df["valor_liquido"].sum()) if "valor_liquido" in df.columns else 0
    tot_vbruto = float(df["valor_bruto"].sum())   if "valor_bruto"   in df.columns else 0

    pct_acerto = round(tot_dev / tot_remet * 100, 1) if tot_remet > 0 else 0

    for col_idx, (db_col, _) in enumerate(cols_excel, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = _fill(COR_TOTAIS)
        cell.font = Font(name="Calibri", bold=True, color=COR_TEXTO_BRANCO, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()

        if col_idx == 1:
            cell.value = f"TOTAIS  |  {len(df)} títulos  |  % Acerto: {pct_acerto:.1f}%"
        elif db_col == "qtde_remessa":
            cell.value = tot_remet
            cell.number_format = "#,##0"
        elif db_col == "qtde_dev_acert":
            cell.value = tot_dev
            cell.number_format = "#,##0"
        elif db_col == "qtde_saldo":
            cell.value = tot_saldo
            cell.number_format = "#,##0"
        elif db_col == "valor_liquido":
            cell.value = tot_vliq
            cell.number_format = 'R$ #,##0.00'
        elif db_col == "valor_bruto":
            cell.value = tot_vbruto
            cell.number_format = 'R$ #,##0.00'

    ws.row_dimensions[3].height = 18

    # ── 7. Linha de cabeçalhos (linha 4) ─────────────────────────────────────
    for col_idx, (db_col, header) in enumerate(cols_excel, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        # Destaque especial para colunas de ação do cliente
        if "Faturar" in header or "Devolver" in header or "Observação" in header:
            cell.fill = _fill(COR_FATURAR)
        else:
            cell.fill = _fill(COR_CABECALHO)
        cell.font = Font(name="Calibri", bold=True, color=COR_TEXTO_BRANCO, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[4].height = 30

    # ── 8. Dados (linhas 5+) ─────────────────────────────────────────────────
    COLS_NUM   = {"qtde_remessa", "qtde_dev_acert", "qtde_saldo"}
    COLS_MONEY = {"valor_liquido", "valor_bruto"}
    COLS_PERC  = {"desconto"}

    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=5):
        is_par = (row_idx % 2 == 0)
        fill_normal = _fill(COR_ZEBRA_PAR if is_par else COR_ZEBRA_IMPAR)
        fill_faturar = _fill(COR_FATURAR_CELL)

        for col_idx, (db_col, header) in enumerate(cols_excel, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _border("hair")

            is_action = db_col is None  # colunas extras (cliente preenche)

            if is_action:
                cell.fill = fill_faturar
                cell.font = _font(size=9, color="1A5E38")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            val = row_data.get(db_col) if hasattr(row_data, "get") else row_data[db_col] if db_col in df.columns else None

            if db_col in COLS_NUM:
                cell.value = int(val) if pd.notna(val) else 0
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.fill = fill_normal
            elif db_col in COLS_MONEY:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = "R$ #,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.fill = fill_normal
            elif db_col in COLS_PERC:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = "0%"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = fill_normal
            elif db_col == "data_emissao":
                if pd.notna(val):
                    try:
                        cell.value = pd.Timestamp(val).to_pydatetime()
                        cell.number_format = "DD/MM/YYYY"
                    except Exception:
                        cell.value = str(val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = fill_normal
            elif db_col in ("titulo", "razao_social", "autor"):
                cell.value = str(val) if pd.notna(val) else ""
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                cell.fill = fill_normal
                cell.font = _font(size=9, bold=(db_col == "titulo"))
            else:
                cell.value = str(val) if pd.notna(val) else ""
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = fill_normal

            cell.font = _font(
                size=9,
                bold=(db_col == "titulo"),
                color="1A5E38" if is_action else "000000"
            )

        ws.row_dimensions[row_idx].height = 14

    # ── 9. Linha de rodapé com soma das colunas de ação ──────────────────────
    footer_row = 5 + len(df)
    for col_idx in range(1, n_cols_total + 1):
        cell = ws.cell(row=footer_row, column=col_idx)
        cell.fill = _fill("F0F4F8")
        cell.border = _border()

    # ── 10. Larguras das colunas ─────────────────────────────────────────────
    LARGURAS = {
        "cod_loja": 10, "cnpj": 20, "codigo_cliente": 8,
        "loja": 5, "cod_gcon": 8, "uf": 5, "razao_social": 28,
        "nf_serie": 16, "data_emissao": 13, "isbn": 15, "cod_barras": 15,
        "titulo": 42, "autor": 22, "desconto": 9, "status_titulo": 16,
        "qtde_remessa": 10, "qtde_dev_acert": 12, "qtde_saldo": 10,
        "valor_liquido": 14, "valor_bruto": 13,
        None: 16,  # colunas de ação
    }

    for col_idx, (db_col, header) in enumerate(cols_excel, start=1):
        key = db_col if db_col in LARGURAS else None
        ws.column_dimensions[get_column_letter(col_idx)].width = LARGURAS.get(key, 12)

    # ── 11. Congela painel no cabeçalho ──────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── 12. Filtro automático nos cabeçalhos ─────────────────────────────────
    ws.auto_filter.ref = f"A4:{get_column_letter(n_cols_total)}{4 + len(df)}"

    # ── 13. Aba de instruções ────────────────────────────────────────────────
    ws_instr = wb.create_sheet("Instruções")
    instrucoes = [
        ("COMO PREENCHER ESTE MAPA DE CONSIGNAÇÃO", True, "1E3A5F", 14),
        ("", False, "000000", 10),
        ("1. Verifique todos os títulos listados na aba 'Mapa de Consignação'.", False, "000000", 10),
        ("2. Para cada título que deseja COMPRAR, preencha a coluna 'Qtde a Faturar ←' com a quantidade desejada.", False, "000000", 10),
        ("3. Para cada título que deseja DEVOLVER, preencha a coluna 'Qtde a Devolver ←'.", False, "000000", 10),
        ("4. Use a coluna 'Observação' para qualquer comentário adicional.", False, "000000", 10),
        ("5. Salve o arquivo e RESPONDA ESTE EMAIL com o arquivo preenchido.", False, "000000", 10),
        ("6. Mantenha o assunto original do email para garantir o rastreamento correto.", False, "1A5E38", 10),
        ("", False, "000000", 10),
        (f"Código de rastreio: {codigo_rastreio}", True, "1E3A5F", 11),
        (f"Referência: {mes_fmt}", False, "5F6368", 10),
        (f"Vendedor responsável: {vendedor_nome}", False, "5F6368", 10),
        (f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", False, "5F6368", 10),
    ]
    ws_instr.column_dimensions["A"].width = 90
    for i, (texto, bold, color, size) in enumerate(instrucoes, start=1):
        cell = ws_instr.cell(row=i, column=1, value=texto)
        cell.font = Font(name="Calibri", bold=bold, color=color, size=size)
        cell.alignment = Alignment(vertical="center")
        ws_instr.row_dimensions[i].height = 18 if bold else 15

    # ── 14. Salva em memória ─────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _fmt_mes(mes_referencia: str) -> str:
    meses = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março",
        "04": "Abril", "05": "Maio", "06": "Junho",
        "07": "Julho", "08": "Agosto", "09": "Setembro",
        "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }
    try:
        ano, mes = mes_referencia.split("-")
        return f"{meses.get(mes, mes)}/{ano}"
    except Exception:
        return mes_referencia or datetime.now().strftime("%m/%Y")


def gerar_mapa_todos_clientes(
    df_saldo_completo: pd.DataFrame,
    clientes_info: list[dict],
    mes_referencia: str,
    vendedor_nome: str,
    cod_gcon: str = None,
) -> dict[str, bytes]:
    """
    Gera mapas Excel para múltiplos clientes.

    Retorna dict: {codigo_cliente: bytes_do_excel}
    """
    resultados = {}
    df = df_saldo_completo.copy()

    if cod_gcon:
        df = df[df.get("cod_gcon", pd.Series(dtype=str)) == cod_gcon]

    for cli_info in clientes_info:
        cod = cli_info.get("codigo_cliente") or cli_info.get("Cod.", "")
        df_cli = df[df["codigo_cliente"] == cod]

        if df_cli.empty:
            continue

        import secrets as sec
        rastreio = sec.token_hex(8).upper()

        try:
            excel_bytes = gerar_mapa_excel(
                df_saldo=df_cli,
                codigo_cliente=cod,
                razao_social=cli_info.get("razao_social", ""),
                cnpj=cli_info.get("cnpj", ""),
                mes_referencia=mes_referencia,
                codigo_rastreio=rastreio,
                vendedor_nome=vendedor_nome,
            )
            resultados[cod] = {"bytes": excel_bytes, "rastreio": rastreio}
        except Exception as e:
            resultados[cod] = {"bytes": None, "erro": str(e)}

    return resultados
